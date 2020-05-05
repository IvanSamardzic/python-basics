import openpyxl as xl #create alias for package
from openpyxl.chart import BarChart, Reference #import modules to grab new prices and make a chart

def process_workbook(filename,sheet_num, chart_cell, discount):   
    #load into wb variable excel file
    wb = xl.load_workbook(filename)
    #into sheet variable add only first sheet 
    sheet = wb['Sheet ' + self.sheet_num]

    #iterate over rows in the excel file 
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * float(self.discount)
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
    
    values = Reference(sheet, 
              min_row = 2,  
              max_row = sheet.max_row,
              min_col = 4, 
              max_col = 4)
    chart = BarChart()
    #add values from values object to the chart
    chart.add_data(values)
    #add chart starting from the chart_cell cell, cell will be in the left top chart corner
    sheet.add_chart(chart, self.chart_cell)
    #append new values and chart into excel file 
    wb.save(filename)